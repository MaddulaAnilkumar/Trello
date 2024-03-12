import openpyxl
class Excel_data():
    def __init__(self,file,sheetname):
        workbook = openpyxl.load_workbook(file)
        self.sheet = workbook.get_sheet_by_name(sheetname)
    def rowcount(self):
        return (self.sheet.max_row)
    def cloumnCount(self):
        return (self.sheet.max_column)
    def readdata(self,rownum,columnnum):
        return self.sheet.cell(row=rownum, column=columnnum).value
    def writedata(self,file,rownum,columnnum, data):
        workbook = openpyxl.load_workbook(file)
        self.sheet.cell(row=rownum, column=columnnum).value = data
        workbook.save(file)



